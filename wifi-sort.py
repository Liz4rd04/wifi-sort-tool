#!/usr/bin/env python3
"""
wifi-sort - Extract WiFi data from Kismet .kismet database files into categorized Excel tabs

Usage:
    wifi-sort input.kismet -o output.xlsx --client patterns.txt [--exclude exclude.txt]

Tabs:
  1. Client-Named     - SSIDs matching --client patterns
  2. Non-Client-Named - Other identified SSIDs (optionally filtered with --exclude)
  3. Unknown Devices  - Devices with no SSID (hard to identify)

Pattern files contain one SSID pattern per line. Supports wildcards:
    testwifi123*      - matches "testwifi123", "testwifi1234", "testwifi12345"
    *xfinity*    - matches "xfinitywifi", "Xfinity Mobile"
"""

import argparse
import fnmatch
import sys
import sqlite3
import json
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: Required packages not installed.", file=sys.stderr)
    print("Run: sudo apt install python3-pandas python3-openpyxl", file=sys.stderr)
    sys.exit(1)


def load_patterns(pattern_file):
    """Load SSID patterns from a file, one per line."""
    if not pattern_file:
        return []
    patterns = []
    with open(pattern_file, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):
                patterns.append(line)
    return patterns


def ssid_matches_patterns(ssid, patterns):
    """Check if an SSID matches any of the given patterns."""
    if not patterns:
        return False
    if ssid is None:
        ssid = ''
    for pattern in patterns:
        if pattern.lower() == '<empty>':
            if ssid == '':
                return True
            continue
        if fnmatch.fnmatch(ssid.lower(), pattern.lower()):
            return True
    return False


def freq_to_channel(freq):
    """Convert frequency in kHz to channel number."""
    if freq is None:
        return None
    freq_mhz = freq / 1000 if freq > 10000 else freq
    
    # 2.4 GHz band
    if 2412 <= freq_mhz <= 2484:
        if freq_mhz == 2484:
            return 14
        return int((freq_mhz - 2407) / 5)
    
    # 5 GHz band
    if 5170 <= freq_mhz <= 5825:
        return int((freq_mhz - 5000) / 5)
    
    # 6 GHz band
    if 5955 <= freq_mhz <= 7115:
        return int((freq_mhz - 5950) / 5)
    
    return None


def extract_kismet_data(db_path, verbose=False):
    """Extract WiFi device data from Kismet database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [row[0] for row in cursor.fetchall()]
    
    if verbose:
        print(f"Tables in database: {tables}")
    
    devices = []
    
    if 'devices' in tables:
        cursor.execute("SELECT device FROM devices")
        rows = cursor.fetchall()
        
        for row in rows:
            try:
                device_json = json.loads(row[0])
                
                phy = device_json.get('kismet.device.base.phyname', '')
                if phy != 'IEEE802.11':
                    continue
                
                mac = device_json.get('kismet.device.base.macaddr', '')
                name = device_json.get('kismet.device.base.name', '')
                dot11 = device_json.get('dot11.device', {})
                
                ssid = ''
                adv_ssids = dot11.get('dot11.device.advertised_ssid_map', [])
                if adv_ssids:
                    for ssid_entry in adv_ssids:
                        ssid = ssid_entry.get('dot11.advertisedssid.ssid', '')
                        if ssid:
                            break
                
                if not ssid:
                    probed = dot11.get('dot11.device.probed_ssid_map', [])
                    if probed:
                        for probe_entry in probed:
                            ssid = probe_entry.get('dot11.probedssid.ssid', '')
                            if ssid:
                                break
                
                freq = device_json.get('kismet.device.base.frequency', 0)
                channel_str = device_json.get('kismet.device.base.channel', '')
                
                channel = None
                if channel_str:
                    try:
                        channel = int(''.join(filter(str.isdigit, channel_str.split('-')[0].split('W')[0][:3])))
                    except:
                        pass
                
                if not channel and freq:
                    channel = freq_to_channel(freq)
                
                signal = device_json.get('kismet.device.base.signal', {})
                rssi = signal.get('kismet.common.signal.last_signal', None)
                min_rssi = signal.get('kismet.common.signal.min_signal', None)
                max_rssi = signal.get('kismet.common.signal.max_signal', None)
                
                location = device_json.get('kismet.device.base.location', {})
                avg_loc = location.get('kismet.common.location.avg_loc', {})
                geopoint = avg_loc.get('kismet.common.location.geopoint', [None, None])
                lat = geopoint[1] if geopoint and len(geopoint) > 1 else None
                lon = geopoint[0] if geopoint and len(geopoint) > 0 else None
                alt = avg_loc.get('kismet.common.location.alt', None)
                
                first_seen = device_json.get('kismet.device.base.first_time', 0)
                last_seen = device_json.get('kismet.device.base.last_time', 0)
                packets_total = device_json.get('kismet.device.base.packets.total', 0)
                packets_data = device_json.get('kismet.device.base.packets.data', 0)
                
                crypt = ''
                if adv_ssids:
                    crypt_set = adv_ssids[0].get('dot11.advertisedssid.crypt_set', 0)
                    crypt_parts = []
                    if crypt_set & 0x02:
                        crypt_parts.append('WEP')
                    if crypt_set & 0x04:
                        crypt_parts.append('WPA')
                    if crypt_set & 0x08:
                        crypt_parts.append('WPA2')
                    if crypt_set & 0x10:
                        crypt_parts.append('WPA3')
                    if crypt_set & 0x200:
                        crypt_parts.append('PSK')
                    if crypt_set & 0x400:
                        crypt_parts.append('Enterprise')
                    crypt = '/'.join(crypt_parts) if crypt_parts else 'Open'
                
                type_str = device_json.get('kismet.device.base.type', 'Unknown')
                manuf = device_json.get('kismet.device.base.manuf', '')
                data_size = device_json.get('kismet.device.base.datasize', 0)
                
                from datetime import datetime
                first_time_str = datetime.fromtimestamp(first_seen).strftime('%Y-%m-%d %H:%M:%S') if first_seen else ''
                last_time_str = datetime.fromtimestamp(last_seen).strftime('%Y-%m-%d %H:%M:%S') if last_seen else ''
                
                devices.append({
                    'MAC': mac,
                    'SSID': ssid,
                    'Type': type_str,
                    'Manufacturer': manuf,
                    'Encryption': crypt,
                    'Channel': channel,
                    'Frequency_MHz': freq / 1000 if freq and freq > 10000 else freq,
                    'RSSI_Last': rssi,
                    'RSSI_Min': min_rssi,
                    'RSSI_Max': max_rssi,
                    'Packets_Total': packets_total,
                    'Packets_Data': packets_data,
                    'Data_Size_Bytes': data_size,
                    'First_Seen': first_time_str,
                    'Last_Seen': last_time_str,
                    'Latitude': lat,
                    'Longitude': lon,
                    'Altitude_m': alt,
                })
                
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                if verbose:
                    print(f"Warning: Could not parse device: {e}")
                continue
    
    conn.close()
    
    if verbose:
        print(f"Extracted {len(devices)} WiFi devices")
    
    return pd.DataFrame(devices)


def write_df_to_sheet(ws, dataframe, sheet_name):
    """Write a DataFrame to an Excel sheet with formatting."""
    ws.title = sheet_name
    
    if dataframe.empty:
        ws.cell(row=1, column=1, value="No matching entries")
        return
    
    headers = list(dataframe.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, row in enumerate(dataframe.values, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    for col_num, header in enumerate(headers, 1):
        max_len = len(str(header))
        for val in dataframe.iloc[:, col_num-1].astype(str).values[:100]:
            max_len = max(max_len, len(val))
        col_letter = chr(64 + col_num) if col_num <= 26 else f"A{chr(64 + col_num - 26)}"
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


def main():
    parser = argparse.ArgumentParser(
        description='Extract Kismet database to Excel with 3 categorized tabs',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Tabs:
  1. Client-Named     - SSIDs matching --client patterns  
  2. Non-Client-Named - Other identified SSIDs (use --exclude to filter out known networks)
  3. Unknown Devices  - Devices with no SSID (hard to identify)

Pattern file format (one pattern per line):
  testwifi123*         Match SSIDs starting with "testwifi123" (case-insensitive)
  *corp*          Match SSIDs containing "corp"
  ssid123*            Match SSIDs starting with "ssid123"
  # comment       Lines starting with # are ignored

Example:
  python3 wifi-sort.py capture.kismet -o sorted.xlsx --client client.txt --exclude known.txt
        """
    )
    
    parser.add_argument('input', help='Input Kismet database file (.kismet)')
    parser.add_argument('-o', '--output', default='output.xlsx', help='Output Excel file (default: output.xlsx)')
    parser.add_argument('--client', required=True, metavar='FILE', help='Pattern file for client SSIDs (tab 1)')
    parser.add_argument('--exclude', metavar='FILE', help='Pattern file for SSIDs to exclude from tab 2 (optional)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Show detailed output')
    
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        print(f"Error: Input file '{args.input}' not found", file=sys.stderr)
        sys.exit(1)
    
    if not Path(args.client).exists():
        print(f"Error: Pattern file '{args.client}' not found", file=sys.stderr)
        sys.exit(1)
    
    if args.exclude and not Path(args.exclude).exists():
        print(f"Error: Exclude file '{args.exclude}' not found", file=sys.stderr)
        sys.exit(1)
    
    client_patterns = load_patterns(args.client)
    exclude_patterns = load_patterns(args.exclude) if args.exclude else []
    
    if not client_patterns:
        print("Error: Client pattern file is empty", file=sys.stderr)
        sys.exit(1)
    
    if args.verbose:
        print(f"Loading Kismet database: {args.input}")
    
    df = extract_kismet_data(args.input, verbose=args.verbose)
    
    if df.empty:
        print("Error: No WiFi devices found in database", file=sys.stderr)
        sys.exit(1)
    
    if args.verbose:
        print(f"Using {len(client_patterns)} client patterns")
        if exclude_patterns:
            print(f"Using {len(exclude_patterns)} exclude patterns")
    
    df['SSID'] = df['SSID'].fillna('')
    
    # Tab 1: Client-Named - matches client patterns
    client_mask = df['SSID'].apply(lambda x: ssid_matches_patterns(x, client_patterns))
    client_df = df[client_mask]
    
    # Tab 3: Unknown Devices - empty SSID
    unknown_mask = df['SSID'] == ''
    unknown_df = df[unknown_mask]
    
    # Tab 2: Non-Client-Named - has SSID, not client, optionally excluding patterns
    has_ssid = df['SSID'] != ''
    not_client = ~client_mask
    
    if exclude_patterns:
        not_excluded = ~df['SSID'].apply(lambda x: ssid_matches_patterns(x, exclude_patterns))
        non_client_df = df[has_ssid & not_client & not_excluded]
    else:
        non_client_df = df[has_ssid & not_client]
    
    # Create workbook
    wb = Workbook()
    
    ws1 = wb.active
    write_df_to_sheet(ws1, client_df, 'Client-Named')
    
    ws2 = wb.create_sheet()
    write_df_to_sheet(ws2, non_client_df, 'Non-Client-Named')
    
    ws3 = wb.create_sheet()
    write_df_to_sheet(ws3, unknown_df, 'Unknown Devices')
    
    wb.save(args.output)
    
    print(f"Created {args.output}:")
    print(f"  Client-Named:     {len(client_df)} devices")
    print(f"  Non-Client-Named: {len(non_client_df)} devices")
    print(f"  Unknown Devices:  {len(unknown_df)} devices")
    
    if args.verbose:
        print(f"\nClient-Named SSIDs:")
        for ssid in sorted(client_df['SSID'].unique()):
            count = len(client_df[client_df['SSID'] == ssid])
            print(f"    {ssid} ({count})")
        
        print(f"\nNon-Client-Named SSIDs:")
        for ssid in sorted(non_client_df['SSID'].unique()):
            count = len(non_client_df[non_client_df['SSID'] == ssid])
            print(f"    {ssid} ({count})")
        
        if exclude_patterns:
            excluded_df = df[has_ssid & not_client & df['SSID'].apply(lambda x: ssid_matches_patterns(x, exclude_patterns))]
            if not excluded_df.empty:
                print(f"\nExcluded SSIDs (not in output):")
                for ssid in sorted(excluded_df['SSID'].unique()):
                    count = len(excluded_df[excluded_df['SSID'] == ssid])
                    print(f"    {ssid} ({count})")


if __name__ == '__main__':
    main()
